from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
import openpyxl

# Load the workbook
wb = Workbook()
ws = wb.active
wb = openpyxl.load_workbook('w.xlsx')
ws = wb.active
rule = ColorScaleRule(start_type='num', start_value=0, start_color='AA0000',
                      mid_type='num', mid_value=0.5, mid_color='FFFF00',
                      end_type='num', end_value=1, end_color='00AA00')


# Get the range of cells in the worksheet
min_cell, max_cell = ws.dimensions.split(':')
all_range = min_cell + ':' + max_cell
# Apply the color scale rule to all cells in the worksheet
ws.conditional_formatting.add(all_range, rule)
# Save the workbook
wb.save('w.xlsx')
print('finish')
# import openpyxl
# from openpyxl.styles import PatternFill
#
# # Load the workbook
# wb = openpyxl.load_workbook('workbook.xlsx')
# ws = wb['Sheet1']
# fill_pattern = PatternFill(patternType='solid',fgColor='C64747')
# ws['A2'].fill = fill_pattern
# wb.save("workbook2.xlsx")
