import argparse
import json
import timeit
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
import pandas as pd
import torch
import clip
import wandb
import yaml
from zero_shot_style.evaluation.evaluation_all import get_gts_data, CLIPScoreRef, CLIPScore, STYLE_CLS
from zero_shot_style.evaluation.evaluation_all import evaluate_single_res
from zero_shot_style.evaluation.pycocoevalcap.bleu.bleu import Bleu
from zero_shot_style.evaluation.pycocoevalcap.rouge.rouge import Rouge

from model.ZeroCLIP import CLIPTextGenerator
from datetime import datetime
import os.path
import csv
from collections import defaultdict
import numpy as np
import pickle
from datetime import datetime
from utils import parser, get_hparams
from evaluate import load
MAX_PERPLEXITY = 500

def get_args():
    #parser = argparse.ArgumentParser() #comment when using, in addition, the arguments from zero_shot_style.utils
    #parser.add_argument('--wandb_mode', type=str, default='disabled', help='disabled, offline, online')
    parser.add_argument("--img_name", type=int, default=0)
    parser.add_argument("--use_all_imgs", type=int, default=0)
    parser.add_argument("--seed", type=int, default=0)
    #parser.add_argument("--lm_model", type=str, default="gpt-2", help="gpt-2 or gpt-neo")
    parser.add_argument("--lm_model", type=str, default="gpt-2", help="gpt-2 or gpt-neo or gpt-j")
    parser.add_argument("--clip_checkpoints", type=str, default="./clip_checkpoints", help="path to CLIP")
    parser.add_argument("--target_seq_length", type=int, default=15)
    parser.add_argument("--data_type", type=str, default='val', choices = ['train', 'val', 'test'])
    parser.add_argument("--cond_text", type=str, default="Image of a")
    parser.add_argument("--cond_text_list", nargs="+", type=str, default=["Image of a"])
    # parser.add_argument("--cond_text_list", nargs="+", type=str, default=["A creative short caption I can generate to describe this image is:",
    #                                                                       "A creative positive short caption I can generate to describe this image is:",
    #                                                                       "A creative negative short caption I can generate to describe this image is:",
    #                                                                       "A creative humoristic short caption I can generate to describe this image is:",
    #                                                                       "A creative romantic short caption I can generate to describe this image is:"])
    # parser.add_argument("--cond_text", type=str, default="")
    parser.add_argument("--cond_text2", type=str, default="")
    parser.add_argument("--reset_context_delta", action="store_true",
                        help="Should we reset the context at each token gen")
    parser.add_argument("--clip_loss_temperature", type=float, default=0.01)
    parser.add_argument("--std_embedding_vectors_positive", type=float, default=0.028914157)
    parser.add_argument("--std_embedding_vectors_negative", type=float, default=0.020412436)
    parser.add_argument("--ce_scale", type=float, default=0.2)
    parser.add_argument("--clip_scale", type=float, default=1)
    parser.add_argument("--text_style_scale", type=float, default=1)
    parser.add_argument("--num_iterations", type=int, default=5)
    parser.add_argument("--stepsize", type=float, default=0.3)
    parser.add_argument("--grad_norm_factor", type=float, default=0.9)
    parser.add_argument("--fusion_factor", type=float, default=0.99)
    parser.add_argument("--repetition_penalty", type=float, default=2)
    parser.add_argument("--end_token", type=str, default=".", help="Token to end text")
    parser.add_argument("--end_factor", type=float, default=1.01, help="Factor to increase end_token")
    parser.add_argument("--forbidden_factor", type=float, default=20, help="Factor to decrease forbidden tokens")
    parser.add_argument("--beam_size", type=int, default=5)

    parser.add_argument("--max_num_of_imgs", type=int, default=2)#1e6)
    parser.add_argument("--evaluationo_metrics", nargs="+",
                        default=['clip_score', 'fluency', 'style_cls'])
                        # default=['bleu', 'rouge', 'clip_score_ref', 'clip_score', 'fluency', 'style_cls'])


    parser.add_argument("--cuda_idx_num", type=str, default="0")
    parser.add_argument("--img_idx_to_start_from", type=int, default=0)

    parser.add_argument('--run_type',
                        default='caption',
                        nargs='?',
                        choices=['caption', 'arithmetics','img_prompt_manipulation'])

    parser.add_argument("--dataset_types", type=str, default='single')
    parser.add_argument("--dataset_names", nargs="+", type=str, default=["senticap"])  # todo: add: "flickrstyle10k"])
    parser.add_argument("--caption_img_dict", nargs="+", type=str, default=[os.path.join(os.path.expanduser('~'),'data','senticap'),
                                                                              os.path.join(os.path.expanduser('~'),
                                                                                           'data', 'flickrstyle10k')],
                        help="Path to images dict for captioning")
    # parser.add_argument("--caption_img_dict", type=str, default=[os.path.join(os.path.expanduser('~'),'data','senticap')],
    #                     help="Path to images dict for captioning")
    #parser.add_argument("--caption_img_dict", type=str, default=[os.path.join(os.path.expanduser('~'), 'data', 'flickrstyle10k')],
    #                    help="Path to images dict for captioning")
    '''
    parser.add_argument("--caption_img_dict", type=str, default=[os.path.join(os.path.expanduser('~'),'data','imgs')],
                        help="Path to images dict for captioning")
    '''
    parser.add_argument("--caption_img_path", type=str, default=os.path.join(os.path.expanduser('~'),'data','imgs','101.jpeg'),
                        help="Path to image for captioning")

    parser.add_argument("--arithmetics_imgs", nargs="+",
                        default=['example_images/arithmetics/woman2.jpg',
                                 'example_images/arithmetics/king2.jpg',
                                 'example_images/arithmetics/man2.jpg'])

    parser.add_argument("--arithmetics_style_imgs", nargs="+",
                        default=['49','50','51','52','53'])
    parser.add_argument("--arithmetics_weights", nargs="+", default=[1, 1, -1])
    parser.add_argument("--use_style_model", action="store_true", default=False)

    args = parser.parse_args()

    return args

def run(config, img_path, text_style_scale,imitate_text_style,desired_style_embedding_vector,desired_style_embedding_std_vector,cuda_idx,title2print,model_path,dataset_type,tmp_text_loss,label,img_dict,debug_tracking,text_generator=None,image_features=None):
    #debug_tracking: debug_tracking[img_path][label][word_num][iteration][module]:<list>
    if text_generator == None:
        text_generator = CLIPTextGenerator(cuda_idx=cuda_idx,model_path = model_path,tmp_text_loss= tmp_text_loss, text_style_scale=text_style_scale, **vars(config))
    if image_features == None:
        image_features = text_generator.get_img_feature([img_path], None)

    # SENTIMENT: added scale parameter
    if imitate_text_style:
        text_style = label
    else:
        text_style = ''
    t1 = timeit.default_timer();
    captions = text_generator.run(image_features, config['cond_text'], config['beam_size'],text_style_scale,text_style,desired_style_embedding_vector,desired_style_embedding_std_vector,dataset_type)
    debug_tracking[img_path][label] = text_generator.get_debug_tracking()
    t2 = timeit.default_timer();
    encoded_captions = [text_generator.clip.encode_text(clip.tokenize(c).to(text_generator.device)) for c in captions]
    encoded_captions = [x / x.norm(dim=-1, keepdim=True) for x in encoded_captions]
    best_clip_idx = (torch.cat(encoded_captions) @ image_features.t()).squeeze().argmax().item()

    print(captions)

    dt_string = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    new_title2print = f'~~~~~~~~\n{dt_string} | Work on img path:' + title2print.split(' | Work on img path:')[1]
    print(new_title2print)

    print('best clip:', config['cond_text'] + captions[best_clip_idx])
    print(f"Time to create caption is: {(t2-t1)/60} minutes = {t2-t1} seconds.")
    img_dict[img_path][dataset_type][text_style_scale][label] = config['cond_text'] + captions[best_clip_idx]
    return config['cond_text'] + captions[best_clip_idx]

def run_arithmetic(text_generator,config,model_path, img_dict_img_arithmetic,base_img,dataset_type, imgs_path, img_weights, cuda_idx,title2print):
    if text_generator == None:
        text_generator = CLIPTextGenerator(cuda_idx=cuda_idx,model_path = model_path, **vars(config))
    # text_generator = CLIPTextGenerator(cuda_idx=cuda_idx, **vars(config))

    image_features = text_generator.get_combined_feature(imgs_path, [], img_weights, None)
    t1 = timeit.default_timer();
    captions = text_generator.run(image_features, config['cond_text'], beam_size=config['beam_size'])
    t2 = timeit.default_timer();

    encoded_captions = [text_generator.clip.encode_text(clip.tokenize(c).to(text_generator.device)) for c in captions]
    encoded_captions = [x / x.norm(dim=-1, keepdim=True) for x in encoded_captions]
    best_clip_idx = (torch.cat(encoded_captions) @ image_features.t()).squeeze().argmax().item()

    print(captions)

    dt_string = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    new_title2print = f'~~~~~~~~\n{dt_string} | Work on img path:' + title2print.split(' | Work on img path:')[1]
    print(new_title2print)

    print('best clip:', config['cond_text'] + captions[best_clip_idx])
    print(f"Time to create caption is: {(t2-t1)/60} minutes = {t2-t1} seconds.")
    img_dict_img_arithmetic[base_img][dataset_type] = config['cond_text'] + captions[best_clip_idx]

    return config['cond_text'] + captions[best_clip_idx]

def run_img_and_prompt_manipulation(config, img_dict_img_arithmetic,base_img,dataset_type, imgs_path, img_weights, cuda_idx,title2print):
    #text_generator = CLIPTextGenerator(**vars(args))
    text_generator = CLIPTextGenerator(cuda_idx=cuda_idx, **vars(config))

    image_features = text_generator.get_combined_feature(imgs_path, [], img_weights, None)
    t1 = timeit.default_timer();
    captions = text_generator.run(image_features, config['cond_text'], beam_size=config['beam_size'])
    t2 = timeit.default_timer();

    encoded_captions = [text_generator.clip.encode_text(clip.tokenize(c).to(text_generator.device)) for c in captions]
    encoded_captions = [x / x.norm(dim=-1, keepdim=True) for x in encoded_captions]
    best_clip_idx = (torch.cat(encoded_captions) @ image_features.t()).squeeze().argmax().item()

    print(captions)

    dt_string = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    new_title2print = f'~~~~~~~~\n{dt_string} | Work on img path:' + title2print.split(' | Work on img path:')[1]
    print(new_title2print)

    print('best clip:', config['cond_text'] + captions[best_clip_idx])
    print(f"Time to create caption is: {(t2-t1)/60} minutes = {t2-t1} seconds.")
    img_dict_img_arithmetic[base_img][dataset_type] = config['cond_text'] + captions[best_clip_idx]

    return config['cond_text'] + captions[best_clip_idx]


# SENTIMENT: writing results to file
def write_results(img_dict):
    with open('results.csv', 'w') as results_file:
        writer = csv.writer(results_file)
        for img in img_dict.keys():    
            writer.writerow([img])
            writer.writerow(['scale/sentiment', 'negative', 'positive', 'neutral','factual'])
            for scale in img_dict[img].keys():
                cur_row = [scale]
                for sentiment in img_dict[img][scale].keys():
                    cur_row.append(img_dict[img][scale][sentiment])
                writer.writerow(cur_row)

def write_results_of_text_style(img_dict, embedding_type,labels,results_dir,dataset_type):
    if not os.path.isdir(results_dir):
        os.makedirs(results_dir)
    tgt_path = os.path.join(results_dir,f'results_{dataset_type}_embedding_type_{embedding_type}.csv')
    print(f'Writing results into: {tgt_path}')
    with open(tgt_path, 'w') as results_file:
        writer = csv.writer(results_file)
        for img in img_dict.keys():
            img_num_str = img.split('/')[-1].split('.j')[0]
            writer.writerow([img_num_str])
            titles = ['scale/label']
            titles.extend(labels)
            writer.writerow(titles)
            for scale in img_dict[img].keys():
                cur_row = [scale]
                for label in img_dict[img][scale].keys():
                    cur_row.append(img_dict[img][scale][label])
                writer.writerow(cur_row)


def write_debug_tracking(results_dir,debug_tracking):
    '''
    create file 'debug_tracking.xlsx' in results_dir
    :param img_dict: 
    :param labels: 
    :param results_dir: 
    :param scales_len: 
    :param debug_tracking: debug_tracking[img_path][label][word_num][iteration][module]:<list>
    :return: 
    '''
    # img_dict[img_path][dataset_type][text_style_scale][label]
    if not os.path.isdir(results_dir):
        os.makedirs(results_dir)
    tgt_debugging_results_path = os.path.join(results_dir, 'debug_tracking.xlsx')
    print(f'Writing results into: {tgt_debugging_results_path}')
    rule = ColorScaleRule(start_type='num', start_value=0, start_color='AA0000',
                          mid_type='num', mid_value=0.5, mid_color='FFFF00',
                          end_type='num', end_value=1, end_color='00AA00')
    # Load the workbook
    wb = Workbook()
    ws = wb.active

    title = ['img_name','label','word_num','iteration_num','module/prob','1','2','3','4','5','6','7','8','9','10','img_path']
    r = 0
    for i in range(len(title)):
        ws.cell(row=r+1, column=i+1).value=title[i]
    r += 1
    for img_path in debug_tracking.keys():
        img_name_str = img_path.split('/')[-1].split('.j')[0]
        for label in debug_tracking[img_path]:
            for word_num in debug_tracking[img_path][label]:
                for iteration_num in debug_tracking[img_path][label][word_num]:
                    for module in debug_tracking[img_path][label][word_num][iteration_num]:
                        row = [img_name_str,label,word_num,iteration_num,module]
                        row.extend(debug_tracking[img_path][label][word_num][iteration_num][module])
                        row.append(img_path)
                        for i in range(len(row)):
                            ws.cell(row=r+1, column=i+1).value = row[i]
                        r += 1

    # Get the range of cells in the worksheet
    min_cell, max_cell = ws.dimensions.split(':')
    all_range = 'F2' + ':' + 'O'+max_cell[1:]
    # Apply the color scale rule to all cells in the worksheet
    ws.conditional_formatting.add(all_range, rule)
    # Save the workbook
    wb.save(tgt_debugging_results_path)
    print(f'Finished to write debug tracking into: {tgt_debugging_results_path}')


def write_results_of_text_style_all_models(img_dict,labels,results_dir,scales_len,tgt_results_path):
    # img_dict[img_path][dataset_type][text_style_scale][label]
    if not os.path.isdir(results_dir):
        os.makedirs(results_dir)
    print(f'Writing results into: {tgt_results_path}')
    with open(tgt_results_path, 'w') as results_file:
        writer = csv.writer(results_file)
        for img in img_dict.keys():
            img_num_str = img.split('/')[-1].split('.j')[0]
            titles0 = ['img_num']
            titles0.extend([img_num_str] * scales_len*len(labels))
            writer.writerow(titles0)
            titles1 = ['label']
            for label in labels:
                titles1.extend([label] * scales_len)
            writer.writerow(titles1)
            titles2 = ['model/scale']
            titles2.extend(list(img_dict[img][list(img_dict[img].keys())[0]].keys()) * len(labels))
            writer.writerow(titles2)
            for model_name in img_dict[img]:
                cur_row = [model_name]
                for scale in img_dict[img][model_name].keys():
                    for label in img_dict[img][model_name][scale].keys():
                        cur_row.append(img_dict[img][model_name][scale][label])
                writer.writerow(cur_row)


def write_results_prompt_manipulation(img_dict,results_dir,tgt_results_path):
    # img_dict[img_path][dataset_type][text_style_scale][label]
    if not os.path.isdir(results_dir):
        os.makedirs(results_dir)
    print(f'Writing results into: {tgt_results_path}')
    writeTitle = True
    with open(tgt_results_path, 'w') as results_file:
        writer = csv.writer(results_file)
        for i,img in enumerate(img_dict.keys()):
            img_num_str = img.split('/')[-1].split('.j')[0]
            cur_row = [img_num_str]
            for model_name in img_dict[img]:
                for scale in img_dict[img][model_name].keys():
                    labels = img_dict[img][model_name][scale].keys()
                    if writeTitle:
                        titles0 = ['img_num\prompt']
                        titles0.extend(labels)
                        writer.writerow(titles0)
                        writeTitle = False
                    for label in labels:
                        cur_row.append(img_dict[img][model_name][scale][label])
            writer.writerow(cur_row)

def write_img_idx_to_name(img_idx_to_name, tgt_results_path):
    print(f'Writing img idx to name: {tgt_results_path}..')
    with open(tgt_results_path, 'w') as results_file:
        writer = csv.writer(results_file)
        for i in img_idx_to_name.keys():
            cur_row = [i, img_idx_to_name[i]]
            writer.writerow(cur_row)
    print(f'Finished to write img idx to name: {tgt_results_path}')


def write_results_image_manipulation(img_dict_img_arithmetic,results_dir,tgt_results_path):
    if not os.path.isdir(results_dir):
        os.makedirs(results_dir)
    print(f'Writing results into: {tgt_results_path}')
    writeTitle = True
    with open(tgt_results_path, 'w') as results_file:
        writer = csv.writer(results_file)
        for i, img in enumerate(img_dict_img_arithmetic.keys()):
            # img_num_str = img.split('/')[-1].split('.j')[0]
            img_num_str = str(img)
            cur_row = [img_num_str]
            styles = img_dict_img_arithmetic[img].keys()
            if writeTitle:
                titles0 = ['img_num\style']
                titles0.extend(styles)
                writer.writerow(titles0)
                writeTitle = False
            for dataset_type in styles:
                cur_row.append(img_dict_img_arithmetic[img][dataset_type])
            writer.writerow(cur_row)


def write_evaluation_results(total_captions,avg_total_score, results_dir, config):
    if not os.path.isdir(results_dir):
        os.makedirs(results_dir)
    tgt_results_path = os.path.join(results_dir,f"avg_total_score={avg_total_score}_LM_loss_scale={config['ce_scale']}_CLIP_loss_scale={config['clip_scale']}_STYLE_loss_scale={config['text_style_scale']}.csv")
    print(f'Writing evaluation results into: {tgt_results_path}')
    with open(tgt_results_path, 'w') as results_file:
        writer = csv.writer(results_file)
        title = ['img_name', 'style', 'caption', 'gt_captions', 'factual_captions', 'avg_style_cls_score', 'avg_clip_score', 'avg_fluency_score', 'avg_total_score', 'ce_scale','clip_scale','text_style_scale','beam_size','num_iterations','img_path']
        writer.writerow(title)
        for i in total_captions:
            cur_row = [i.get_img_name(), i.get_style(), i.get_caption_text(),i.get_gt_caption_text(), i.get_factual_captions(), i.get_style_cls_score(),i.get_clip_score(), i.get_fluency_score(), i.get_total_score(),config['ce_scale'],config['clip_scale'],config['text_style_scale'],config['beam_size'],config['num_iterations'], i.get_img_path()]
            writer.writerow(cur_row)

def get_title2print(caption_img_path, dataset_type, label, text_style_scale,config):
    dt_string = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    title2print = f"~~~~~~~~\n{dt_string} | Work on img path: {caption_img_path} with:" \
                  f"\nresults dir= *** {config['experiment_dir']} ***" \
                  f"\ndataset_type= *** {dataset_type} ***" \
                  f"\nstyle of: *** {label} ***\ntext_style_scale= *** {text_style_scale} ***" \
                  f"\n~~~~~~~~"
    return title2print


def get_full_path_of_stylized_images(data_dir, i):
    if os.path.isfile(os.path.join(data_dir, 'stylized_images',
                                   str(i) + ".jpeg")):
        return os.path.join(data_dir, 'stylized_images',
                            str(i) + ".jpeg")
    elif os.path.isfile(os.path.join(data_dir, 'stylized_images',
                                     str(i) + ".jpg")):
        return os.path.join(data_dir, 'stylized_images',
                            str(i) + ".jpg")
    elif os.path.isfile(os.path.join(data_dir, 'stylized_images',
                                   str(i) + ".png")):
        return os.path.join(data_dir, 'stylized_images',
                            str(i) + ".png")
    else:
        return None


def calculate_avg_score(clip_score, fluency_score,style_cls_score=None):
    if style_cls_score:
        avg_total_score = 3*(style_cls_score*clip_score*fluency_score)/(style_cls_score+clip_score+fluency_score)
    else:
        avg_total_score = 3*(clip_score*fluency_score)/(clip_score+fluency_score)
    return avg_total_score


def get_img_full_path(base_path, i):
    if os.path.isfile(os.path.join(base_path, 'data', 'imgs',
                                   str(i) + ".jpeg")):
        return os.path.join(base_path, 'data', 'imgs',
                            str(i) + ".jpeg")
    elif os.path.isfile(os.path.join(base_path, 'data', 'imgs',
                                     str(i) + ".jpg")):
        return os.path.join(base_path, 'data', 'imgs',
                            str(i) + ".jpg")
    elif os.path.isfile(os.path.join(base_path, 'data', 'imgs',
                                   str(i) + ".png")):
        return os.path.join(base_path, 'data', 'imgs',
                            str(i) + ".png")
    else:
        return None


class Caption:
    def __init__(self, img_name, style, res_caption_text,  gt_caption_text, img_path,classification,clip_score,fluency,avg_total_score, factual_captions,style_cls_score=None):
        self.img_name = img_name
        self.style = style
        self.res_caption_text = res_caption_text
        self.gt_caption_text = gt_caption_text
        self.img_path = img_path
        self.perplexity = fluency
        self.avg_style_cls_score = style_cls_score
        self.avg_clip_score = clip_score
        self.avg_fluency_score = fluency
        self.classification = classification
        self.avg_total_score = avg_total_score
        self.factual_captions = factual_captions

    def get_img_name(self):
        return self.img_name

    def get_style(self):
        return self.style

    def get_caption_text(self):
        return self.res_caption_text

    def get_gt_caption_text(self):
        return self.gt_caption_text

    def get_factual_captions(self):
        return self.factual_captions

    def get_img_path(self):
        return self.img_path

    def get_classification(self):
        return self.classification

    def get_style_cls_score(self):
        return self.avg_style_cls_score

    def get_clip_score(self):
        return self.avg_clip_score

    def get_fluency_score(self):
        return self.avg_fluency_score

    def get_total_score(self):
        return self.avg_total_score

    def set_classification(self,classification):
        self.classification = classification

    def set_style_cls_score(self,avg_style_cls_score):
        self.avg_style_cls_score = avg_style_cls_score

    def set_clip_score(self,avg_clip_score):
        self.avg_clip_score = avg_clip_score

    def set_fluency_score(self,avg_fluency_score):
        self.avg_fluency_score = avg_fluency_score

    def set_total_score(self,avg_total_score):
        self.avg_total_score = avg_total_score

class Fluency:
    def __init__(self):
        self.model_id = 'gpt2'
        self.perplexity = load("perplexity", module_type="measurement")
        self.tests = []
        self.img_names = []
        self.styles = []

    def add_test(self, res, img_name, style):
        self.tests.append(res)
        self.img_names.append(img_name)
        self.styles.append(style)

    def compute_score(self):
        results = self.perplexity.compute(data=self.tests, model_id=self.model_id, add_start_token=False)
        perplexities = {}
        for i,p in enumerate(results['perplexities']):
            if self.img_names[i] not in perplexities:
                perplexities[self.img_names[i]] = {}
            perplexities[self.img_names[i]][self.styles[i]] = 1-np.min([p,MAX_PERPLEXITY])/MAX_PERPLEXITY
        total_avg_perplexity = 1-np.min([results['mean_perplexity'],MAX_PERPLEXITY])/MAX_PERPLEXITY
        return perplexities, total_avg_perplexity


def get_table_for_wandb(data_list):
    data = [[x, y] for (x, y) in zip(data_list, list(range(len(data_list))))]
    table = wandb.Table(data=data, columns=["x", "y"])
    return table



def get_factual_captions(factual_file_path_list):
    '''
    get dict of all factual captions
    :param factual_file_path_list: list of file path to factual data
    :return: factual_captions:dict:key=img_name,value=list of factual captions
    '''
    print("create dict of factual captions...")
    factual_captions = {}
    finish_im_ids = []
    for factual_file_path in factual_file_path_list:
        data = json.load(open(factual_file_path, "r"))
        for d in data['annotations']:
            if d['image_id'] in finish_im_ids:
                continue
            if d['image_id'] not in factual_captions:
                factual_captions[d['image_id']] = [d['caption']]
            else:
                factual_captions[d['image_id']].append(d['caption'])
        finish_im_ids = list(factual_captions.keys())
    return factual_captions


def main():
    args = get_args()
    config = get_hparams(args)
    os.environ["CUDA_VISIBLE_DEVICES"] = config['cuda_idx_num']

    checkpoints_dir = os.path.join(os.path.expanduser('~'), 'checkpoints')
    data_dir = os.path.join(os.path.expanduser('~'), 'data')
    factual_captions_path = os.path.join(data_dir,'source','coco','factual_captions.pkl')

    text_style_scale_list = [config['text_style_scale']]
    if not config['use_style_model']:
        config['text_style_scale'] = 0
    text_to_imitate_list = ["bla"]#["Happy", "Love", "angry", "hungry", "I love you!!!", " I hate you and I want to kill you",
                            #"Let's set a meeting at work", "I angry and I love", "The government is good"]
    imitate_text_style = False

    dataset_type_list = config["data_name"]
    if config['dataset_types'] =='all':
        dataset_type_list = ['senticap', 'flickrstyle10k']

    imgs_dataset_type_dict = {49: 'factual', 50: 'positive', 51: 'negative', 52: 'humor', 53: 'romantic'}
    # label2prompt = {"factual": "Image of a", "positive": "The beautiful image of a",
    #                 "ngeative": "The disturbing image of a"}
    if len(config["cond_text_list"])==1:
        label2prompt = {"factual": config["cond_text_list"][0], "positive": config["cond_text_list"][0],
                        "negative": config["cond_text_list"][0]}
    else:
        label2prompt = {"factual": config["cond_text_list"][0], "positive": config["cond_text_list"][1],
                        "negative": config["cond_text_list"][2]}


    if config['run_type'] == 'img_prompt_manipulation':
        prompt2idx_img_style = {config['cond_text_list'][0]: 49, config['cond_text_list'][1]: 50,
                                config['cond_text_list'][2]: 51}

    with open(factual_captions_path,'rb') as f:
        factual_captions = pickle.load(f)

    wandb.init(project='StylizedZeroCap',
               config=config,
               resume=config['resume'],
               id=config['run_id'],
               mode=config['wandb_mode'], #disabled, offline, online'
               tags=config['tags'])

    # handle sweep training names
    config['training_name'] = f'{wandb.run.id}-{wandb.run.name}'

    config['experiment_dir'] = f'{os.path.expanduser("~")}/experiments/stylized_zero_cap_experiments/5_3_23/{config["training_name"]}'
    wandb.config.update(config, allow_val_change=True)

    labels_dict_idxs = {'positive': 0, 'negative':1, 'humor': 0, 'romantic':1}

    txt_cls_model_paths = {'senticap': os.path.join(os.path.expanduser('~'),'checkpoints','best_models','senticap','pos_neg_best_text_style_classification_model.pth'),
                           'flickrstyle10k': os.path.join(os.path.expanduser('~'),'checkpoints','best_models','humor_romantic_best_text_style_classification_model.pth')}



    print(f'saving experiment outputs in {os.path.abspath(config["experiment_dir"])}')

    if not os.path.exists(config['experiment_dir']):
        os.makedirs(config['experiment_dir'])

    print('------------------------------------------------------------------------------------------------------')
    print('Training config:')
    for k, v in config.items():
        print(f'{k}: {v}')
    print('------------------------------------------------------------------------------------------------------')
    with open(os.path.join(config['experiment_dir'], 'config.pkl'), 'wb') as f:
        pickle.dump(config, f)
    with open(os.path.join(config['experiment_dir'], 'config.yaml'), 'w') as f:
        yaml.dump(config, f, default_flow_style=False)
    print('saved experiment config in ', os.path.join(config['experiment_dir'], 'config.pkl'))
    results_dir = config['experiment_dir']

    cur_time = datetime.now().strftime("%H_%M_%S__%d_%m_%Y")
    print(f'Cur time is: {cur_time}')
    img_dict = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: ""))))
    img_dict_img_arithmetic = defaultdict(lambda: defaultdict(lambda: "")) #img_path,dataset_type
    tmp_text_loss = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: "")))
    debug_tracking = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: []))) #word_num:iteration:module:list
    model_path = os.path.join(checkpoints_dir, 'best_models',
                              config['best_model_name'])


    #for imitate_text_style in [False]:
    if imitate_text_style:
        classes_type = "sentences"
    else:
        classes_type = "source"

    if config['calc_fluency']:
        fluency_obj = Fluency()

    data_set_path = {}
    txt_cls_model_paths_to_load = {}
    for dataset_name in config['dataset_names']:
        txt_cls_model_paths_to_load[dataset_name] = txt_cls_model_paths[dataset_name]
        data_set_path[dataset_name] = os.path.join(data_dir, dataset_name, 'annotations', config['data_type']+'.pkl')
    gts_per_data_set = get_gts_data(data_set_path,factual_captions)

    if not config['debug_mac']:
        text_generator = CLIPTextGenerator(cuda_idx=config['cuda_idx_num'], model_path=model_path, tmp_text_loss=tmp_text_loss,
                                           **config)
    else:
        text_generator = None

    evaluation_obj = {}
    for metric in config['evaluationo_metrics']:
        if metric=='bleu':
            evaluation_obj['bleu'] = Bleu(n=4)
        if metric=='rouge':
            evaluation_obj['rouge'] = Rouge()
        if metric == 'clip_score_ref':
            evaluation_obj['clip_score_ref'] = CLIPScoreRef(text_generator)
        if metric == 'clip_score':
            evaluation_obj['clip_score'] = CLIPScore(text_generator)
        if metric == 'fluency':
            evaluation_obj['fluency'] = Fluency()
        if metric == 'style_cls':
            evaluation_obj['style_cls'] = STYLE_CLS(txt_cls_model_paths_to_load, data_dir, config['cuda_idx_num'], labels_dict_idxs) #todo:change handling dataset_type qs list

    # take list of images
    imgs_to_test = []
    print(f"config['max_num_of_imgs']: {config['max_num_of_imgs']}")
    for dataset_type in dataset_type_list:
        config['caption_img_dict'] = os.path.join(os.path.expanduser('~'), 'data', dataset_type)
        for i,im in enumerate(os.listdir(os.path.join(config['caption_img_dict'],'images',config['data_type']))):
            if i >= config['max_num_of_imgs'] and config['max_num_of_imgs']>0:
                break
            if ('.jpg' or '.jpeg' or '.png') not in im:
                continue
            imgs_to_test.append(os.path.join(config['caption_img_dict'],'images',config['data_type'],im))
    print(f"***There is {len(imgs_to_test)} images to test.***")

    # go over all images
    evaluation_results = {}
    for img_path_idx, img_path in enumerate(imgs_to_test):  # img_path_list:
        # if int(img_path.split('.')[0].split('/')[-1]) == 429063:
        #     print(f'img_path_idx={img_path_idx}')
        wandb.log({'test/img_idx': img_path_idx})
        print(f"Img num = {img_path_idx}")
        if not config['debug_mac']:
            image_features = text_generator.get_img_feature([img_path], None)
        else:
            image_features = None
        if img_path_idx < config['img_idx_to_start_from']:
            continue
        img_name = int(img_path.split('/')[-1].split('.')[0])
        config['caption_img_path'] = img_path
        evaluation_results[img_name] = {'img_path': img_path}
        if not os.path.isfile(config['caption_img_path']):
            continue
        mean_embedding_vec_path = os.path.join(checkpoints_dir, 'best_models',
                                               config['mean_vec_emb_file'])
        std_embedding_vec_path = os.path.join(checkpoints_dir, 'best_models',
                                               config['std_vec_emb_file'])
        # desired_labels_list = config['desired_labels']
        for dataset_type in dataset_type_list:
            if config['use_style_model']:
                with open(mean_embedding_vec_path, 'rb') as fp:
                    embedding_vectors_to_load = pickle.load(fp)
                with open(std_embedding_vec_path, 'rb') as fp:
                    std_embedding_vectors_to_load = pickle.load(fp)
                std_embedding_vectors_to_load={'positive': config['std_embedding_vectors_positive'], 'negative': config['std_embedding_vectors_negative']}

                desired_labels_list = list(embedding_vectors_to_load.keys())
            else:
                desired_labels_list = config['desired_labels']
                # desired_labels_list = [prompt]
            if imitate_text_style:
                desired_labels_list = text_to_imitate_list
            use_style_model = config['use_style_model']
            source_text_style_scale_list = text_style_scale_list
            source_clip_scale = config['clip_scale']
            source_ce_scale = config['ce_scale']
            source_text_style_scale = config['text_style_scale']
            source_beam_size = config['beam_size']
            source_num_iterations = config['num_iterations']
            tgt_results_path = os.path.join(results_dir, f'results_all_models_{classes_type}_classes_{cur_time}.csv')
            for label in desired_labels_list:
                prompt = label2prompt[label]
                config['cond_text'] = prompt
                if label=='factual':
                    config['use_style_model'] = False
                    text_style_scale_list = [0]
                    config['clip_scale'] = 1
                    config['ce_scale'] = 0.2
                    config['beam_size'] = 5
                    config['num_iterations'] = 5
                    config['text_style_scale'] = 0
                    text_generator.set_params(config['ce_scale'], config['clip_scale'], config['text_style_scale'],
                                              config['beam_size'], config['num_iterations'])
                else:
                    if config['use_style_model']:
                        config['use_style_model'] = use_style_model
                        text_style_scale_list = source_text_style_scale_list
                        config['clip_scale'] = source_clip_scale
                        config['ce_scale'] = source_ce_scale
                        config['beam_size'] = source_beam_size
                        config['num_iterations'] = source_num_iterations
                        config['text_style_scale'] = source_text_style_scale
                    else:
                        text_style_scale_list = [0]
                        config['clip_scale'] = 1
                        config['ce_scale'] = 0.2
                        config['text_style_scale'] = 0
                        config['beam_size'] = 5
                        config['num_iterations'] = 5
                    text_generator.set_params(config['ce_scale'], config['clip_scale'], config['text_style_scale'],
                                              config['beam_size'], config['num_iterations'])

                evaluation_results[img_name][label] = {}
                desired_style_embedding_vector = ''
                desired_style_embedding_std_vector = ''
                if not imitate_text_style:
                    if config['use_style_model']:
                        desired_style_embedding_vector = embedding_vectors_to_load[label]
                        desired_style_embedding_std_vector = std_embedding_vectors_to_load[label]
                print(f"Img num = {img_path_idx}")
                if config['debug_mac']:
                    evaluation_results[img_name][label] = {'res': 'bla'}
                    continue
                if config['run_type'] == 'caption':
                    title2print = get_title2print(config['caption_img_path'], dataset_type, label,
                                                  config['text_style_scale'], config)
                    print(title2print)
                    best_caption = run(config, config['caption_img_path'],
                        config['text_style_scale'], imitate_text_style, desired_style_embedding_vector, desired_style_embedding_std_vector,
                        config['cuda_idx_num'], title2print, model_path, dataset_type,tmp_text_loss,label,img_dict,debug_tracking,text_generator,image_features)
                    config['use_style_model'] = use_style_model
                    if not config['use_style_model']:
                        write_results_prompt_manipulation(img_dict, results_dir, tgt_results_path)
                    if config['use_style_model']:
                        write_results_of_text_style_all_models(img_dict, desired_labels_list,
                                                           results_dir, 1, tgt_results_path)
                        write_debug_tracking(results_dir, debug_tracking)
                        fluency_obj.add_test(best_caption, img_name, label)

                elif config['run_type'] == 'arithmetics':
                    config['cond_text'] = ""
                    #none arithmetic
                    title2print = get_title2print(config['caption_img_path'], dataset_type, 'factual',
                                                  config['text_style_scale'],config)
                    print(title2print)
                    config['arithmetics_imgs'] = [config['caption_img_path'], config['caption_img_path'], config['caption_img_path']]
                    best_caption = run_arithmetic(text_generator,config,model_path, img_dict_img_arithmetic, img_name,
                                   'factual', imgs_path=config['arithmetics_imgs'],
                                   img_weights=[1, 0, 0], cuda_idx=config['cuda_idx_num'],title2print = title2print)

                    write_results_image_manipulation(img_dict_img_arithmetic, results_dir, tgt_results_path)

                    config['arithmetics_weights'] = [float(x) for x in config['arithmetics_weights']]
                    factual_img_style = get_full_path_of_stylized_images(data_dir, config['arithmetics_style_imgs'][0])
                    for idx, v in enumerate(config['arithmetics_style_imgs'][1:]):
                        img_style = get_full_path_of_stylized_images(data_dir, v)
                        config['arithmetics_imgs'] = [config['caption_img_path'], factual_img_style, img_style]

                        title2print = get_title2print(config['caption_img_path'], dataset_type, imgs_dataset_type_dict[int(v)],
                                                      config['text_style_scale'], config)

                        best_caption = run_arithmetic(text_generator,config,model_path,img_dict_img_arithmetic,img_name,imgs_dataset_type_dict[int(v)], imgs_path=config['arithmetics_imgs'],
                                       img_weights=config['arithmetics_weights'], cuda_idx=config['cuda_idx_num'],title2print = title2print)

                        write_results_image_manipulation(img_dict_img_arithmetic, results_dir, tgt_results_path)

                elif config['run_type'] == 'img_prompt_manipulation':
                    config['arithmetics_weights'] = [float(x) for x in config['arithmetics_weights']]
                    factual_img_style = get_full_path_of_stylized_images(data_dir, config['arithmetics_style_imgs'][0])
                    config['cond_text'] = prompt
                    v = prompt2idx_img_style[prompt]
                    img_style = get_full_path_of_stylized_images(data_dir, v)
                    config['arithmetics_imgs'] = [config['caption_img_path'], factual_img_style, img_style]

                    title2print = get_title2print(config['caption_img_path'], dataset_type, imgs_dataset_type_dict[int(v)],
                                                  config['text_style_scale'], config)

                    best_caption = run_arithmetic(text_generator,config,model_path,img_dict_img_arithmetic,img_name,imgs_dataset_type_dict[int(v)], imgs_path=config['arithmetics_imgs'],
                                   img_weights=config['arithmetics_weights'], cuda_idx=config['cuda_idx_num'],title2print = title2print)

                    write_results_image_manipulation(img_dict_img_arithmetic, results_dir, tgt_results_path)

                else:
                    raise Exception('run_type must be caption or arithmetics!')
                evaluation_results[img_name][label]['res'] = best_caption


    #add gt to evaluation_results dict
    txt_cls_model_paths = {'senticap': os.path.join(os.path.expanduser('~'),'checkpoints','best_models','pos_neg_best_text_style_classification_model.pth'),
                           'flickrstyle10k': os.path.join(os.path.expanduser('~'),'checkpoints','best_models','humor_romantic_best_text_style_classification_model.pth')}
    #calc evaluation
    print("Calc evaluation of the results...")
    #calc perplexity
    perplexities,mean_perplexity = fluency_obj.compute_score()
    # for debug
    # results = fluency_obj.perplexity.compute(data=fluency_obj.tests, model_id=fluency_obj.model_id, add_start_token=False)

    # factual_captions = get_factual_captions(factual_file_path_list)

    style_cls_scores = []
    clip_scores = []
    fluency_scores = []
    total_captions = []
    total_score_and_text = []
    total_res_text = []
    total_gt_text = []
    avg_total_scores = []
    for img_name in evaluation_results:
        # for label in evaluation_results[img_name]:
        for label in list(evaluation_results[img_name].keys()): #todo:debug
            if label == 'img_path':
                continue
            evaluation_results[img_name][label]['gt'] = gts_per_data_set[dataset_type][img_name][label] #todo: handle style type
            evaluation_results[img_name][label]['scores'] = evaluate_single_res(
                evaluation_results[img_name][label]['res'], evaluation_results[img_name][label]['gt'],
                evaluation_results[img_name]['img_path'], label, dataset_name, config['evaluationo_metrics'],
                evaluation_obj)


            clip_score = evaluation_results[img_name][label]['scores']['clip_score']
            fluency_score = perplexities[img_name][label]
            if 'style_cls' in evaluation_results[img_name][label]['scores']:
                style_cls_score = evaluation_results[img_name][label]['scores']['style_cls']
                avg_total_score = calculate_avg_score(clip_score, fluency_score, style_cls_score)
                style_cls_scores.append(style_cls_score)
            else:
                style_cls_score='None'
                avg_total_score = calculate_avg_score(clip_score, fluency_score)

            clip_scores.append(clip_score)
            fluency_scores.append(fluency_score)
            avg_total_scores.append(avg_total_score)
            res_text = evaluation_results[img_name][label]['res']
            gt_text = evaluation_results[img_name][label]['gt']

            total_res_text.append(res_text)
            total_gt_text.append(gt_text)
            total_captions.append(Caption(img_name, label, evaluation_results[img_name][label]['res'],evaluation_results[img_name][label]['gt'], evaluation_results[img_name]['img_path'],label,clip_score,fluency_score,avg_total_score,factual_captions[img_name],style_cls_score))
            #todo: check if to unindent


    clip_scores_table = get_table_for_wandb(clip_scores)
    fluency_scores_table = get_table_for_wandb(fluency_scores)
    avg_total_scores_table = get_table_for_wandb(avg_total_scores)

    if 'style_cls' in evaluation_results[img_name][label]['scores']:
        style_cls_scores_table = get_table_for_wandb(style_cls_scores)
        style_cls_scores_data = [[x, y] for (x, y) in zip(style_cls_scores, list(range(len(style_cls_scores))))]
        style_cls_scores_table = wandb.Table(data=style_cls_scores_data, columns=["x", "y"])
        wandb.log({'details_evaluation/style_cls_score': style_cls_scores_table})

    clip_scores_data = [[x, y] for (x, y) in zip(clip_scores, list(range(len(clip_scores))))]
    clip_scores_table = wandb.Table(data=clip_scores_data, columns=["x", "y"])
    wandb.log({'details_evaluation/clip_scores': clip_scores_table})

    fluency_scores_data = [[x, y] for (x, y) in zip(fluency_scores, list(range(len(fluency_scores))))]
    fluency_scores_table = wandb.Table(data=fluency_scores_data, columns=["x", "y"])
    wandb.log({'details_evaluation/fluency_score': fluency_scores_table})

    avg_total_scores_data = [[x, y] for (x, y) in zip(avg_total_scores, list(range(len(avg_total_scores))))]
    avg_total_scores_table = wandb.Table(data=avg_total_scores_data, columns=["x", "y"])
    wandb.log({'details_evaluation/avg_total_score': avg_total_scores_table})


    # wandb.log({'details_evaluation/style_cls_score': style_cls_scores_table,
    #            'details_evaluation/clip_score': clip_scores_table,
    #            'details_evaluation/fluency_score': fluency_scores_table,
    #            'details_evaluation/avg_total_score': avg_total_scores_table})

    # avg_total_scores = [0.2,0.3,0.4]
    total_score_and_text = pd.concat(
        [pd.DataFrame({'avg_total_score': avg_total_scores}, index=list(range(len(avg_total_scores)))),
         pd.DataFrame({'total_res_text': total_res_text}, index=list(range(len(total_res_text)))),
         pd.DataFrame({'total_gt_text': total_gt_text}, index=list(range(len(total_gt_text))))], axis=1)
    wandb.log({'details_evaluation/total_score_text': total_score_and_text})

    # fluency_scores.append(evaluation_results[img_name][label]['scores']['fluency'])
    avg_clip_score = np.mean(clip_scores)
    avg_fluency_score = mean_perplexity

    if style_cls_score!='None':
        avg_style_cls_score = np.mean(style_cls_scores)
        final_avg_total_score = calculate_avg_score(avg_clip_score, avg_fluency_score, avg_style_cls_score)
    else:
        avg_style_cls_score = 0
        final_avg_total_score = calculate_avg_score(avg_clip_score,avg_fluency_score)

    print("*****************************")
    print("*****************************")
    print("*****************************")
    if style_cls_score!='None':
        print(f'style_cls_scores={style_cls_scores}')
    print(f'clip_scores={clip_scores},\nfluency_scores={fluency_scores}')
    print(f'final_avg_total_score={final_avg_total_score}')
    print("*****************************")
    print("*****************************")
    print("*****************************")
    wandb.log({'evaluation/mean_style_cls_scores':avg_style_cls_score,
               'evaluation/mean_clip_scores':avg_clip_score,
               'evaluation/mean_fluency_scores':avg_fluency_score,
               'evaluation/final_avg_total_score':final_avg_total_score})

    write_evaluation_results(total_captions,final_avg_total_score, results_dir, config)
    print('Finish of program!')

if __name__ == "__main__":
    main()
