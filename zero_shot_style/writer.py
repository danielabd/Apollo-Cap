import os
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
import csv

# SENTIMENT: writing results to file
def write_results(img_dict):
    with open('results.csv', 'w') as results_file:
        writer = csv.writer(results_file)
        for img in img_dict.keys():
            writer.writerow([img])
            writer.writerow(['scale/sentiment', 'negative', 'positive', 'neutral','factual'])
            for scale in img_dict[img].keys():
                cur_row = [scale]
                for sentiment in img_dict[img][scale].keys():
                    cur_row.append(img_dict[img][scale][sentiment])
                writer.writerow(cur_row)

def write_results_of_text_style(img_dict, embedding_type,labels,results_dir,dataset_type):
    if not os.path.isdir(results_dir):
        os.makedirs(results_dir)
    tgt_path = os.path.join(results_dir,f'results_{dataset_type}_embedding_type_{embedding_type}.csv')
    print(f'Writing results into: {tgt_path}')
    with open(tgt_path, 'w') as results_file:
        writer = csv.writer(results_file)
        for img in img_dict.keys():
            img_num_str = img.split('/')[-1].split('.j')[0]
            writer.writerow([img_num_str])
            titles = ['scale/label']
            titles.extend(labels)
            writer.writerow(titles)
            for scale in img_dict[img].keys():
                cur_row = [scale]
                for label in img_dict[img][scale].keys():
                    cur_row.append(img_dict[img][scale][label])
                writer.writerow(cur_row)


def write_debug_tracking(results_dir,debug_tracking):
    '''
    create file 'debug_tracking.xlsx' in results_dir
    :param img_dict:
    :param labels:
    :param results_dir:
    :param scales_len:
    :param debug_tracking: debug_tracking[img_path][label][word_num][iteration][module]:<list>
    :return:
    '''
    # img_dict[img_path][dataset_type][text_style_scale][label]
    if not os.path.isdir(results_dir):
        os.makedirs(results_dir)
    tgt_debugging_results_path = os.path.join(results_dir, 'debug_tracking.xlsx')
    print(f'Writing results into: {tgt_debugging_results_path}')
    rule = ColorScaleRule(start_type='num', start_value=0, start_color='AA0000',
                          mid_type='num', mid_value=0.5, mid_color='FFFF00',
                          end_type='num', end_value=1, end_color='00AA00')
    # Load the workbook
    wb = Workbook()
    ws = wb.active

    title = ['img_name','label','word_num','iteration_num','module/prob','1','2','3','4','5','6','7','8','9','10','img_path']
    r = 0
    for i in range(len(title)):
        ws.cell(row=r+1, column=i+1).value=title[i]
    r += 1
    for img_path in debug_tracking.keys():
        img_name_str = img_path.split('/')[-1].split('.j')[0]
        for label in debug_tracking[img_path]:
            for word_num in debug_tracking[img_path][label]:
                for iteration_num in debug_tracking[img_path][label][word_num]:
                    for module in debug_tracking[img_path][label][word_num][iteration_num]:
                        row = [img_name_str,label,word_num,iteration_num,module]
                        row.extend(debug_tracking[img_path][label][word_num][iteration_num][module])
                        row.append(img_path)
                        for i in range(len(row)):
                            ws.cell(row=r+1, column=i+1).value = row[i]
                        r += 1

    # Get the range of cells in the worksheet
    min_cell, max_cell = ws.dimensions.split(':')
    all_range = 'F2' + ':' + 'O'+max_cell[1:]
    # Apply the color scale rule to all cells in the worksheet
    ws.conditional_formatting.add(all_range, rule)
    # Save the workbook
    wb.save(tgt_debugging_results_path)
    print(f'Finished to write debug tracking into: {tgt_debugging_results_path}')


def write_results_of_text_style_all_models(img_dict,labels,results_dir,scales_len,tgt_results_path):
    # img_dict[img_path][dataset_type][text_style_scale][label]
    if not os.path.isdir(results_dir):
        os.makedirs(results_dir)
    print(f'Writing results into: {tgt_results_path}')
    with open(tgt_results_path, 'w') as results_file:
        writer = csv.writer(results_file)
        for img in img_dict.keys():
            img_num_str = img.split('/')[-1].split('.j')[0]
            titles0 = ['img_num']
            titles0.extend([img_num_str] * scales_len*len(labels))
            writer.writerow(titles0)
            titles1 = ['label']
            for label in labels:
                titles1.extend([label] * scales_len)
            writer.writerow(titles1)
            titles2 = ['model/scale']
            titles2.extend(list(img_dict[img][list(img_dict[img].keys())[0]].keys()) * len(labels))
            writer.writerow(titles2)
            for model_name in img_dict[img]:
                cur_row = [model_name]
                for scale in img_dict[img][model_name].keys():
                    for label in img_dict[img][model_name][scale].keys():
                        cur_row.append(img_dict[img][model_name][scale][label])
                writer.writerow(cur_row)

def write_caption_results(img_dict,results_dir,tgt_results_path):
    # img_dict[img_path][dataset_type][text_style_scale][label]
    if not os.path.isdir(results_dir):
        os.makedirs(results_dir)
    print(f'Writing results into: {tgt_results_path}')
    writeTitle = True
    with open(tgt_results_path, 'w') as results_file:
        writer = csv.writer(results_file)
        for i,img in enumerate(img_dict.keys()):
            img_num_str = img.split('/')[-1].split('.j')[0]
            cur_row = [img_num_str]
            for model_name in img_dict[img]:
                for scale in img_dict[img][model_name].keys():
                    labels = img_dict[img][model_name][scale].keys()
                    if writeTitle:
                        titles0 = ['img_num']
                        titles0.extend(labels)
                        writer.writerow(titles0)
                        writeTitle = False
                    for label in labels:
                        cur_row.append(img_dict[img][model_name][scale][label])
            writer.writerow(cur_row)

def write_img_idx_to_name(img_idx_to_name, tgt_results_path):
    print(f'Writing img idx to name: {tgt_results_path}..')
    with open(tgt_results_path, 'w') as results_file:
        writer = csv.writer(results_file)
        for i in img_idx_to_name.keys():
            cur_row = [i, img_idx_to_name[i]]
            writer.writerow(cur_row)
    print(f'Finished to write img idx to name: {tgt_results_path}')


def write_results_image_manipulation(img_dict_img_arithmetic,results_dir,tgt_results_path):
    if not os.path.isdir(results_dir):
        os.makedirs(results_dir)
    print(f'Writing results into: {tgt_results_path}')
    writeTitle = True
    with open(tgt_results_path, 'w') as results_file:
        writer = csv.writer(results_file)
        for i, img in enumerate(img_dict_img_arithmetic.keys()):
            # img_num_str = img.split('/')[-1].split('.j')[0]
            img_num_str = str(img)
            cur_row = [img_num_str]
            styles = img_dict_img_arithmetic[img].keys()
            if writeTitle:
                titles0 = ['img_num\style']
                titles0.extend(styles)
                writer.writerow(titles0)
                writeTitle = False
            for dataset_type in styles:
                cur_row.append(img_dict_img_arithmetic[img][dataset_type])
            writer.writerow(cur_row)


def write_evaluation_results(total_captions,avg_total_score, results_dir, config):
    if not os.path.isdir(results_dir):
        os.makedirs(results_dir)
    tgt_results_path = os.path.join(results_dir,f"avg_total_score={avg_total_score}_LM_loss_scale={config['ce_scale']}_CLIP_loss_scale={config['clip_scale']}_STYLE_loss_scale={config['text_style_scale']}.csv")
    print(f'Writing evaluation results into: {tgt_results_path}')
    with open(tgt_results_path, 'w') as results_file:
        writer = csv.writer(results_file)
        title = ['img_name', 'style', 'caption', 'gt_captions', 'factual_captions', 'avg_style_cls_score', 'avg_clip_score', 'avg_fluency_score', 'avg_total_score', 'ce_scale','clip_scale','text_style_scale','beam_size','num_iterations','img_path']
        writer.writerow(title)
        for i in total_captions:
            cur_row = [i.get_img_name(), i.get_style(), i.get_caption_text(),i.get_gt_caption_text(), i.get_factual_captions(), i.get_style_cls_score(),i.get_clip_score(), i.get_fluency_score(), i.get_total_score(),config['ce_scale'],config['clip_scale'],config['text_style_scale'],config['beam_size'],config['num_iterations'], i.get_img_path()]
            writer.writerow(cur_row)
